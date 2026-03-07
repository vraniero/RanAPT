from pathlib import Path

MAX_CHARS_PER_FILE = 100_000


def _table_to_markdown(table: list[list]) -> str:
    """Convert a pdfplumber table (list of rows) to a markdown table."""
    if not table or not table[0]:
        return ""
    # Clean cells
    rows = []
    for row in table:
        rows.append([str(cell).strip() if cell else "" for cell in row])

    col_count = max(len(r) for r in rows)
    # Pad short rows
    for r in rows:
        while len(r) < col_count:
            r.append("")

    lines = []
    # Header row
    lines.append("| " + " | ".join(rows[0]) + " |")
    lines.append("| " + " | ".join("---" for _ in rows[0]) + " |")
    # Data rows
    for row in rows[1:]:
        lines.append("| " + " | ".join(row) + " |")
    return "\n".join(lines)


def extract_text_from_pdf(file_path: Path) -> str:
    """Extract text and tables from a PDF using pdfplumber, page-by-page.

    Tables are extracted separately and rendered as markdown tables so that
    Claude can parse numeric data accurately. Non-table text is extracted
    normally alongside.
    """
    try:
        import pdfplumber
    except ImportError:
        return f"[pdfplumber not installed — cannot extract {file_path.name}]"

    parts: list[str] = []
    total_chars = 0

    try:
        with pdfplumber.open(str(file_path)) as pdf:
            for page_num, page in enumerate(pdf.pages, start=1):
                if total_chars >= MAX_CHARS_PER_FILE:
                    parts.append(f"\n[Truncated at {MAX_CHARS_PER_FILE} chars]")
                    break

                parts.append(f"--- Page {page_num} ---")

                # Extract tables first
                tables = page.extract_tables() or []
                table_md_parts: list[str] = []
                for table in tables:
                    if not table:
                        continue
                    md = _table_to_markdown(table)
                    if md:
                        table_md_parts.append(md)

                # Extract full page text (includes text that's inside tables too)
                page_text = page.extract_text() or ""

                if table_md_parts:
                    # Include the plain text for context, then the structured tables
                    if page_text.strip():
                        parts.append(page_text)
                        total_chars += len(page_text)
                    parts.append("\n[Structured tables from this page]:\n")
                    for tmd in table_md_parts:
                        remaining = MAX_CHARS_PER_FILE - total_chars
                        if remaining <= 0:
                            break
                        tmd = tmd[:remaining]
                        parts.append(tmd)
                        total_chars += len(tmd)
                else:
                    remaining = MAX_CHARS_PER_FILE - total_chars
                    page_text = page_text[:remaining]
                    parts.append(page_text)
                    total_chars += len(page_text)

    except Exception as e:
        return f"[Error reading {file_path.name}: {e}]"

    return "\n".join(parts)


def extract_text(file_path: Path) -> str:
    """Extract text from any supported file type."""
    ext = file_path.suffix.lower()

    if ext == ".pdf":
        return extract_text_from_pdf(file_path)

    if ext == ".txt":
        try:
            text = file_path.read_text(encoding="utf-8", errors="replace")
            return text[:MAX_CHARS_PER_FILE]
        except Exception as e:
            return f"[Error reading {file_path.name}: {e}]"

    if ext == ".csv":
        try:
            import csv
            import io
            raw = file_path.read_text(encoding="utf-8", errors="replace")[:MAX_CHARS_PER_FILE]
            reader = csv.reader(io.StringIO(raw))
            all_rows = [row for row in reader if any(c.strip() for c in row)]
            if not all_rows:
                return "[Empty CSV]"
            col_count = max(len(r) for r in all_rows)
            for r in all_rows:
                while len(r) < col_count:
                    r.append("")
            lines: list[str] = []
            lines.append("| " + " | ".join(all_rows[0]) + " |")
            lines.append("| " + " | ".join("---" for _ in all_rows[0]) + " |")
            for row in all_rows[1:]:
                lines.append("| " + " | ".join(row) + " |")
            return "\n".join(lines)
        except Exception as e:
            return f"[Error reading {file_path.name}: {e}]"

    if ext in (".xlsx", ".xls"):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(str(file_path), read_only=True, data_only=True)
            parts: list[str] = []
            total = 0
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                parts.append(f"=== Sheet: {sheet_name} ===\n")

                # Collect all rows, then render as markdown table
                all_rows: list[list[str]] = []
                for row in ws.iter_rows(values_only=True):
                    cells = [str(c) if c is not None else "" for c in row]
                    # Skip completely empty rows
                    if not any(c.strip() for c in cells):
                        continue
                    all_rows.append(cells)
                    total += sum(len(c) for c in cells)
                    if total >= MAX_CHARS_PER_FILE:
                        all_rows.append(["[Truncated]"])
                        break

                if all_rows:
                    # Normalize column count
                    col_count = max(len(r) for r in all_rows)
                    for r in all_rows:
                        while len(r) < col_count:
                            r.append("")

                    # First row as header
                    parts.append("| " + " | ".join(all_rows[0]) + " |")
                    parts.append("| " + " | ".join("---" for _ in all_rows[0]) + " |")
                    for row in all_rows[1:]:
                        parts.append("| " + " | ".join(row) + " |")
                    parts.append("")
                else:
                    parts.append("[Empty sheet]\n")

                if total >= MAX_CHARS_PER_FILE:
                    break
            return "\n".join(parts)
        except Exception as e:
            return f"[Error reading Excel {file_path.name}: {e}]"

    return f"[Unsupported file type: {ext}]"


def build_context_message(files: list[tuple[str, Path]]) -> str:
    """Build a concatenated message with file headers for Claude API."""
    parts: list[str] = []
    for name, path in files:
        parts.append(f"=== FILE: {name} ===")
        parts.append(extract_text(path))
        parts.append("")
    return "\n".join(parts)


def build_context_from_parsed(parsed_files: list) -> str:
    """Build context message from pre-parsed file objects (avoids re-extraction).

    Each item should have .name, .text, and .file_type attributes.
    """
    parts: list[str] = []
    for pf in parsed_files:
        if pf.file_type == "image" or not pf.text:
            continue
        parts.append(f"=== FILE: {pf.name} ===")
        parts.append(pf.text)
        parts.append("")
    return "\n".join(parts)
